from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "All Codes"

# ── colour palette ──────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", start_color="1F4E79")   # dark blue
CAT_FILL   = PatternFill("solid", start_color="BDD7EE")   # light blue
ALT_FILL   = PatternFill("solid", start_color="EBF3FB")   # very light blue
WHITE_FILL = PatternFill("solid", start_color="FFFFFF")

HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CAT_FONT   = Font(name="Arial", bold=True, color="1F4E79", size=10)
BODY_FONT  = Font(name="Arial", size=10)
CODE_FONT  = Font(name="Arial", bold=True, size=10)

thin = Side(style="thin", color="AAAAAA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_border():
    med = Side(style="medium", color="FFFFFF")
    return Border(left=med, right=med, top=med, bottom=med)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── data ────────────────────────────────────────────────────────────────────
# (category, code, description, fee_dollars, notes)
# fee=None → not listed on cheat sheet
rows = [
    # ED ASSESSMENT
    ("ED Assessment","H101","Minor assessment – Day (8am–5pm)",23,""),
    ("ED Assessment","H102","Comprehensive assessment – Day (8am–5pm)",57,""),
    ("ED Assessment","H103","Multiple assessment – Day (8am–5pm)",47,""),
    ("ED Assessment","H104","Reassessment – Day (8am–5pm)",23,">2h, change mgmt, not d/c or referral; max 3/pt/day"),
    ("ED Assessment","H131","Minor assessment – Eve Mon–Thu (5pm–midnight)",30,""),
    ("ED Assessment","H132","Comprehensive assessment – Eve Mon–Thu (5pm–midnight)",75,""),
    ("ED Assessment","H133","Multiple assessment – Eve Mon–Thu (5pm–midnight)",62,""),
    ("ED Assessment","H134","Reassessment – Eve Mon–Thu (5pm–midnight)",30,">2h, change mgmt; max 3/pt/day"),
    ("ED Assessment","H121","Minor assessment – Night (midnight–8am)",40,""),
    ("ED Assessment","H122","Comprehensive assessment – Night (midnight–8am)",100,""),
    ("ED Assessment","H123","Multiple assessment – Night (midnight–8am)",81,""),
    ("ED Assessment","H124","Reassessment – Night (midnight–8am)",40,">2h, change mgmt; max 3/pt/day"),
    ("ED Assessment","H151","Minor assessment – Wkd/Hol (8am–midnight)",34,"Incl Fri eve 5pm–midnight"),
    ("ED Assessment","H152","Comprehensive assessment – Wkd/Hol (8am–midnight)",86,"Incl Fri eve 5pm–midnight"),
    ("ED Assessment","H153","Multiple assessment – Wkd/Hol (8am–midnight)",70,"Incl Fri eve 5pm–midnight"),
    ("ED Assessment","H154","Reassessment – Wkd/Hol (8am–midnight)",34,"Incl Fri eve 5pm–midnight; max 3/pt/day"),
    # Premiums
    ("Premiums","H112","Night premium (midnight–8am)",51,"Per patient visit"),
    ("Premiums","H113","Wkd/Hol + Fri eve premium",32,"Per patient visit; incl Fri eve 5pm–midnight"),
    ("Premiums","H114","Weekday eve premium (Mon–Thu)",24,"Per patient visit; NEW April 2026"),
    # Consultations
    ("Consultations","H055","Consultation to ED – FRCPC",126,"Must record referring MD billing # or contact info"),
    ("Consultations","H065","Consultation to ED – CFPC",96,"Must record referring MD billing # or contact info"),
    ("Consultations","K734","Phone consult – referring physician",31,"≥10 min; cannot lead to follow-up"),
    ("Consultations","K735","Phone consult – consultant physician",40,"≥10 min; cannot lead to follow-up"),
    ("Consultations","K736","CritiCall telephone consultation – referring physician",32,"Exact fee $32.45"),
    # On-Call
    ("On-Call","H960","On-call travel – Weekday Day",37,"Max 2/day"),
    ("On-Call","H962","On-call travel – Weekday Eve",37,"Max 2"),
    ("On-Call","H963","On-call travel – Wkd/Hol",37,"Max 4"),
    ("On-Call","H964","On-call travel – Night",37,"Unlimited"),
    ("On-Call","H980","On-call 1st patient – Weekday Day",20,"Max 5"),
    ("On-Call","H984","On-call 1st patient – Weekday Eve",60,"Max 5"),
    ("On-Call","H968","On-call 1st patient – Wkd/Hol",75,"Max 10"),
    ("On-Call","H986","On-call 1st patient – Night",100,"Unlimited"),
    ("On-Call","H981","On-call addl patient – Weekday Day",20,"Max 5"),
    ("On-Call","H985","On-call addl patient – Weekday Eve",60,"Max 5"),
    ("On-Call","H989","On-call addl patient – Wkd/Hol",75,"Max 10"),
    ("On-Call","H987","On-call addl patient – Night",100,"Unlimited"),
    # Counselling / Forms
    ("Counselling / Forms","K005","Mental health counselling",63,"Min 20 min; document mental health dx + times"),
    ("Counselling / Forms","K028","STD/Needlestick management counselling",63,"Min 20 min; document times"),
    ("Counselling / Forms","K015","Counsel family re: death/catastrophe",63,"Min 20 min; document times"),
    ("Counselling / Forms","K623","Form 1 – psychiatric assessment",105,""),
    ("Counselling / Forms","K070","Homecare application",32,""),
    ("Counselling / Forms","K035","Medical conditions report",36,"Mandatory report to MOT; exact fee $36.25 (MOH Jan 2025)"),
    ("Counselling / Forms","A771","Death certificate",34,""),
    ("Counselling / Forms","A777","Pronouncement + death certificate",34,""),
    ("Counselling / Forms","H105","Interim admission orders",29,""),
    ("Counselling / Forms","A320","MAID consultation",174,"NEW April 2026; exact fee $174.25"),
    # Critical Care – Imminent Life-Threatening
    ("Critical Care – Imminent","G521","Imminent life-threatening: 1st 15 min (1st MD)",112,"Arrest/Trauma; document times; cannot see other pts simultaneously; exact fee $111.80 (MOH Jan 2025)"),
    ("Critical Care – Imminent","G523","Imminent life-threatening: 2nd 15 min (1st MD)",58,"Arrest/Trauma; exact fee $57.65 (MOH Jan 2025)"),
    ("Critical Care – Imminent","G522","Imminent life-threatening: each addl 15 min (max 6)",38,"Arrest/Trauma; bill for 2nd/3rd MD too; exact fee $38.00 (MOH Jan 2025)"),
    ("Critical Care – Imminent","E420","Trauma ISS > 15 premium",None,"Add to critical care codes"),
    # Critical Care – Potential Life-Threatening
    ("Critical Care – Potential","G395","Potential life-threatening: 1st 15 min",57,"Other resuscitation; document times; exact fee $57.45 (MOH Jan 2025)"),
    ("Critical Care – Potential","G391","Potential life-threatening: each addl 15 min (max 7)",31,"Bill for 2nd/3rd MD too; exact fee $30.60 (MOH Jan 2025)"),
    # Other CC Procedures
    ("Critical Care – Other Procedures","Z437","Cardioversion (max 3/day)",92,"Allowed with G395 but NOT G521"),
    ("Critical Care – Other Procedures","G269","Central line",31,""),
    ("Critical Care – Other Procedures","G211","Intubation",38,"Allowed with G395 but NOT G521; exact fee $38.35 (MOH Jan 2025)"),
    ("Critical Care – Other Procedures","Z341","Chest tube",70,""),
    ("Critical Care – Other Procedures","Z611","Catheter",9,"Exact fee $8.55 (MOH Jan 2025)"),
    ("Critical Care – Other Procedures","H100","Ultrasound",20,""),
    # Anaesthesia modifiers
    ("Anaesthesia Modifiers","E009","Patient < 1 year old",None,"Add C4 units; 1 unit = $15.96 (April 2026)"),
    ("Anaesthesia Modifiers","E019","Patient 1–8 years old",None,"Add C2 units"),
    ("Anaesthesia Modifiers","E007","Patient 70–79 years old",None,"Add C1 unit"),
    ("Anaesthesia Modifiers","E018","Patient > 80 years old",None,"Add C3 units"),
    ("Anaesthesia Modifiers","E011","Prone position",None,"Add C4 units"),
    ("Anaesthesia Modifiers","E024","Sitting position",None,"Add C4 units"),
    ("Anaesthesia Modifiers","E022","ASA III",None,"Add C2 units"),
    ("Anaesthesia Modifiers","E017","ASA IV",None,"Add C10 units"),
    ("Anaesthesia Modifiers","E016","ASA V",None,"Add C20 units"),
    ("Anaesthesia Modifiers","E020","ASA E (emergency)",None,"Add C4 units"),
    ("Anaesthesia Modifiers","E010","BMI > 40",None,"Add C2 units"),
    ("Anaesthesia Modifiers","E400","Eve/Wkd/Hol time qualifier",None,"3rd line; C = sum of steps 2+3 units"),
    ("Anaesthesia Modifiers","E401","Overnight time qualifier",None,"3rd line; C = sum of steps 2+3 units"),
    ("Anaesthesia Modifiers","E413","Procedure premium – Overnight",None,"+40% on anaesthesia"),
    ("Anaesthesia Modifiers","E412","Procedure premium – Eve/Wkd/Hol",None,"+20% on anaesthesia"),
    # Sutures
    ("Sutures","Z154","Face suture 0–5 cm",39,"Adhesive strips/glue = 50% of fee"),
    ("Sutures","Z177","Face suture 5.1–10 cm",78,""),
    ("Sutures","Z190","Face suture 10.1–15 cm",111,""),
    ("Sutures","Z192","Face suture > 15 cm",170,""),
    ("Sutures","Z176","Body suture 0–5 cm",22,""),
    ("Sutures","Z175","Body suture 5.1–10 cm",39,""),
    ("Sutures","Z179","Body suture 10.1–15 cm",55,""),
    ("Sutures","Z191","Body suture > 15 cm",85,""),
    # Complex Lacerations
    ("Complex Lacerations","R024","Acute earlobe repair",101,"> 20 min; document times"),
    ("Complex Lacerations","Z189","Complex laceration zone 1",92,"> 20 min; document times"),
    ("Complex Lacerations","Z187","Complex face laceration",92,"> 20 min; document times"),
    ("Complex Lacerations","R606","Phalanx amputation repair",161,""),
    ("Complex Lacerations","Z188","Complex non-face laceration",92,"> 20 min; document times"),
    ("Complex Lacerations","P036","Vaginal laceration repair",54,""),
    ("Complex Lacerations","R525","Muscle & skin repair",89,""),
    ("Complex Lacerations","R578","Extensor tendon repair",164,""),
    # Routine Procedures
    ("Routine Procedures","J149C","U/S guided procedure",47,""),
    ("Routine Procedures","G269","Central line",31,""),
    ("Routine Procedures","G224","Nerve block – ring/small (provides analgesia > 4h)",16,""),
    ("Routine Procedures","G268","Arterial line",31,""),
    ("Routine Procedures","G060","Major nerve block (provides analgesia > 4h)",55,""),
    ("Routine Procedures","G061","Minor nerve block (provides analgesia > 4h)",30,""),
    ("Routine Procedures","G260","Major plexus block – 3-in-1 (provides analgesia > 4h)",80,""),
    ("Routine Procedures","G231","Dental block",34,""),
    ("Routine Procedures","G370","Knee aspiration",20,""),
    ("Routine Procedures","G328","Joint aspiration (not knee)",40,""),
    ("Routine Procedures","E446","U/S guided aspiration (addl)",30,"Add-on code"),
    ("Routine Procedures","Z080","Wound/ulcer debridement",20,""),
    ("Routine Procedures","Z331","Diagnostic thoracentesis",32,""),
    ("Routine Procedures","Z332","Therapeutic thoracentesis",59,""),
    ("Routine Procedures","Z590","Diagnostic paracentesis",31,""),
    ("Routine Procedures","Z591","Therapeutic paracentesis",58,""),
    ("Routine Procedures","G356","NG tube insertion",34,""),
    ("Routine Procedures","Z520","G tube change",11,""),
    ("Routine Procedures","Z538","Hernia reduction",25,""),
    ("Routine Procedures","Z543","Proctoscopy",9,""),
    ("Routine Procedures","Z110","Subungual hematoma drainage",17,""),
    ("Routine Procedures","Z608","Manual Foley declogging",59,""),
    ("Routine Procedures","Z128","Nail removal",33,""),
    ("Routine Procedures","A920","Misoprostol (including assessment)",161,""),
    ("Routine Procedures","Z130","Nail removal + cautery",63,""),
    ("Routine Procedures","S756","Removal of POC from OS",112,""),
    ("Routine Procedures","G313","EKG",4,"Must be discharged patient"),
    ("Routine Procedures","Z804","Lumbar puncture",68,""),
    ("Routine Procedures","Z399","Endoscopy",93,""),
    ("Routine Procedures","L810","Description of CSF",25,"Exact fee $25.00 (MOH Jan 2025)"),
    ("Routine Procedures","G435","Tonometry",5,"Exact fee $5.10 (MOH Jan 2025)"),
    ("Routine Procedures","Z432","Exam under anaesthesia",54,""),
    ("Routine Procedures","Z756","Rectal disimpaction / fecal disimpaction",46,"Was $36.80 → $46.00 April 2026"),
    ("Routine Procedures","H264","ED pelvic exam with speculum",12,"NEW April 2026"),
    # Incision & Drainage
    ("Incision & Drainage","Z101","Abscess I&D – one abscess (LA)",28,""),
    ("Incision & Drainage","Z226","Elbow bursa I&D",97,""),
    ("Incision & Drainage","Z102","Abscess I&D – one abscess (GA)",49,""),
    ("Incision & Drainage","Z506","Oral abscess I&D",51,""),
    ("Incision & Drainage","Z173","Abscess I&D – two abscesses (LA)",33,""),
    ("Incision & Drainage","Z714","Bartholin's cyst I&D (LA)",17,""),
    ("Incision & Drainage","Z174","Abscess I&D – 3+ abscesses (LA)",45,""),
    ("Incision & Drainage","Z715","Bartholin's cyst I&D (GA)",51,""),
    ("Incision & Drainage","Z104","Perianal abscess I&D (LA)",33,"Was $20.10 → $33.25 (+65%) April 2026"),
    ("Incision & Drainage","Z140","Breast abscess I&D (LA)",33,""),
    ("Incision & Drainage","Z105","Perianal abscess I&D (GA)",72,""),
    ("Incision & Drainage","Z740","Breast abscess I&D (GA)",75,""),
    ("Incision & Drainage","Z106","Pilonidal abscess I&D (LA)",49,""),
    ("Incision & Drainage","Z227","IM abscess/hematoma I&D",101,""),
    ("Incision & Drainage","Z107","Pilonidal abscess I&D (GA)",118,""),
    ("Incision & Drainage","Z510","Peritonsillar abscess (PTA) drainage",91,""),
    ("Incision & Drainage","Z545","Hemorrhoid I&D",25,""),
    ("Incision & Drainage","E318","Pinna hematoma drainage",92,""),
    # Foreign Bodies
    ("Foreign Bodies","Z114","Foreign body removal – skin (LA)",25,""),
    ("Foreign Bodies","Z915","Foreign body removal – ear (LA)",11,""),
    ("Foreign Bodies","Z115","Foreign body removal – skin (GA)",89,""),
    ("Foreign Bodies","Z866","Foreign body removal – ear (GA)",51,""),
    ("Foreign Bodies","R517","Foreign body removal – muscle",108,""),
    ("Foreign Bodies","Z311","Foreign body removal – nose (LA)",11,""),
    ("Foreign Bodies","Z541","Foreign body removal – rectum (GA)",58,""),
    ("Foreign Bodies","Z312","Foreign body removal – nose (GA)",51,""),
    ("Foreign Bodies","Z847","Foreign body removal – eye",33,""),
    ("Foreign Bodies","Z432A","Foreign body removal – vagina (GA)",54,""),
    ("Foreign Bodies","S547","Foreign body removal – urethra",171,""),
    ("Foreign Bodies","E023C","Foreign body removal – vagina with sedation",None,"6 anaesthesia units"),
    # ENT
    ("ENT","G420","Ear syringe/irrigation",11,""),
    ("ENT","G403","Epley manoeuvre",21,""),
    ("ENT","Z315","Anterior nasal pack",15,""),
    ("ENT","Z321","Direct laryngoscopy",61,""),
    ("ENT","Z316","Posterior nasal pack",36,""),
    ("ENT","Z322","Direct laryngoscopy with FB removal",106,""),
    ("ENT","Z314","Nasal cautery",12,""),
    ("ENT","Z324","Indirect laryngoscopy with FB removal",45,""),
    ("ENT","F136","Nasal fracture reduction",102,""),
    ("ENT","D062","TMJ reduction",52,""),
    # Fractures – Upper Extremity
    ("Fractures – Upper Extremity","F004","Phalanx finger fracture – closed, no reduction",49,"Exact fee $49.20 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F005","Phalanx finger fracture – closed, with reduction",99,"Ortho f/u = 75%; Exact fee $99.25 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","E558","Phalanx finger – each addl fracture with reduction",22,"Exact fee $22.25 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F007","Phalanx finger fracture – open, with reduction",298,"Exact fee $298.45 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F008","Metacarpal fracture – no reduction",49,"Exact fee $49.20 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F009","Metacarpal fracture – with reduction",99,"Exact fee $99.25 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F012","Bennett's fracture – no reduction",49,"Exact fee $49.20 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F013","Bennett's fracture – with reduction",120,"Exact fee $119.80 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F018","Scaphoid fracture – no reduction",49,"Exact fee $49.20 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F102","Other carpal fracture – no reduction",49,"Exact fee $49.20 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F016","Other carpal fracture – with reduction",115,"Exact fee $115.10 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F027","Colles/Smith fracture – no reduction",68,"Exact fee $67.75 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F028","Colles/Smith fracture – with reduction",109,"Exact fee $109.45 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F046","Colles/Smith fracture – with reduction + sedation",149,"Exact fee $149.35 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F031","Radius or ulna fracture – no reduction",81,"Exact fee $81.30 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F032","Radius or ulna fracture – with reduction",118,"Exact fee $117.85 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F034","Olecranon fracture – no reduction",126,"Exact fee $126.25 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F035","Olecranon fracture – with reduction",129,"Exact fee $129.00 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F029","Epicondyle fracture (no cast) – no reduction",68,"Exact fee $67.75 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F037","Epicondyle fracture (no cast) – with reduction",126,"Exact fee $126.25 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F042","Humeral shaft fracture (no cast) – no reduction",68,"Exact fee $67.80 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F043","Humeral shaft fracture (no cast) – with reduction",148,"Exact fee $147.60 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F053","Humeral neck fracture (no cast) – no reduction",68,"Exact fee $67.80 (MOH Jan 2025)"),
    ("Fractures – Upper Extremity","F054","Humeral neck fracture (no cast) – with reduction",134,"Exact fee $133.60 (MOH Jan 2025)"),
    # Fractures – Lower Extremity
    ("Fractures – Lower Extremity","F134","Pelvis fracture with pelvic binder – with reduction",442,"No fee for no-reduction code; Exact fee $442.45 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F095","Femur fracture – with reduction",407,"No fee for no-reduction code; Exact fee $407.35 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F085","Patella fracture (no cast) – no reduction",68,"Exact fee $67.75 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F078","Tibia ± fibula fracture – no reduction",116,"Exact fee $115.95 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F079","Tibia ± fibula fracture – with reduction",180,"Exact fee $180.05 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F082","Fibula fracture – no reduction",68,"Exact fee $67.75 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F083","Fibula fracture – with reduction",101,"Exact fee $101.25 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F074","Ankle fracture – no reduction",68,"Exact fee $67.75 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F075","Ankle fracture – with reduction",145,"Exact fee $144.80 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F104","Ankle fracture with plafond burst – with reduction",242,"Exact fee $242.25 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F070","Calcaneus fracture – no reduction",97,"Exact fee $97.35 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F071","Calcaneus fracture – with reduction",161,"Exact fee $161.45 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F061","Metatarsal fracture (no cast) – no reduction",49,"Exact fee $49.20 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F062","Metatarsal fracture (with cast) – no reduction",68,"Exact fee $67.75 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F063","Metatarsal fracture (with cast) – with reduction",98,"Exact fee $98.35 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F066","Tarsal bone fracture – no reduction",98,"Exact fee $98.10 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F067","Tarsal bone fracture – with reduction",165,"Exact fee $165.20 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F056","Phalanx toe fracture – no reduction",49,"Exact fee $49.20 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","F058","Phalanx toe fracture – with reduction",72,"Exact fee $72.35 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","E560","Phalanx toe – each addl, no reduction",12,"Exact fee $12.05 (MOH Jan 2025)"),
    ("Fractures – Lower Extremity","E561","Phalanx toe – each addl, with reduction",15,"Exact fee $14.90 (MOH Jan 2025)"),
    # Dislocation Reductions
    ("Dislocation Reductions","D001","Phalanx finger dislocation reduction",58,"Exact fee $57.50 (MOH Jan 2025)"),
    ("Dislocation Reductions","E576","Phalanx finger – each addl dislocation",10,"Exact fee $10.25 (MOH Jan 2025)"),
    ("Dislocation Reductions","D003","Phalanx finger dislocation reduction – open",197,"Exact fee $196.50 (MOH Jan 2025)"),
    ("Dislocation Reductions","D004","Metacarpophalangeal dislocation reduction",58,"Exact fee $57.50 (MOH Jan 2025)"),
    ("Dislocation Reductions","D007","Carpal dislocation reduction",128,"Exact fee $128.05 (MOH Jan 2025)"),
    ("Dislocation Reductions","D012","Pulled elbow reduction",39,"Exact fee $39.00 (MOH Jan 2025)"),
    ("Dislocation Reductions","D009","Elbow dislocation reduction",84,"Exact fee $84.45 (MOH Jan 2025)"),
    ("Dislocation Reductions","D015","Shoulder dislocation reduction – no sedation",49,"Exact fee $49.20 (MOH Jan 2025)"),
    ("Dislocation Reductions","D016","Shoulder dislocation reduction – with sedation",111,"Exact fee $111.40 (MOH Jan 2025)"),
    ("Dislocation Reductions","D014","A/C or S/C joint – no reduction",68,"Exact fee $67.80 (MOH Jan 2025)"),
    ("Dislocation Reductions","D042","Hip dislocation reduction",268,"Exact fee $268.25 (MOH Jan 2025)"),
    ("Dislocation Reductions","D038","Knee dislocation reduction (no cast)",208,"Exact fee $207.90 (MOH Jan 2025)"),
    ("Dislocation Reductions","D040","Patella dislocation – no sedation, no cast",62,"Exact fee $62.20 (MOH Jan 2025)"),
    ("Dislocation Reductions","D031","Patella dislocation – with sedation, no cast",97,"Exact fee $97.35 (MOH Jan 2025)"),
    ("Dislocation Reductions","D035","Ankle dislocation reduction",111,"Exact fee $111.35 (MOH Jan 2025)"),
    ("Dislocation Reductions","D026","Tarso-metatarsal dislocation reduction (no cast)",148,"Exact fee $147.60 (MOH Jan 2025)"),
    ("Dislocation Reductions","D030","Metatarsophalangeal dislocation reduction",58,"Exact fee $57.50 (MOH Jan 2025)"),
    ("Dislocation Reductions","D027","Interphalangeal toe dislocation reduction",58,"Exact fee $57.50 (MOH Jan 2025)"),
    # Splints & Casts
    ("Splints & Casts","Z201","Finger splint/cast",10,"Only if no F or D code billed; Exact fee $10.25 (MOH Jan 2025)"),
    ("Splints & Casts","Z202","Hand splint/cast",15,"Only if no F or D code billed; Exact fee $14.90 (MOH Jan 2025)"),
    ("Splints & Casts","Z203","Arm/forearm/wrist splint/cast",24,"Only if no F or D code billed; Exact fee $24.10 (MOH Jan 2025)"),
    ("Splints & Casts","Z213","Below-knee splint/cast",24,"Only if no F or D code billed; Exact fee $24.10 (MOH Jan 2025)"),
    ("Splints & Casts","Z211","Long leg splint/cast",29,"Only if no F or D code billed; Exact fee $28.80 (MOH Jan 2025)"),
    ("Splints & Casts","Z204","Cast removal",10,"Exact fee $10.25 (MOH Jan 2025)"),
]

# ── write header ─────────────────────────────────────────────────────────────
headers = ["Code", "Description", "Fee ($)", "Category", "Notes"]
col_widths = [10, 52, 10, 30, 50]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = CENTER
    cell.border = hdr_border()
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

# ── write data rows ───────────────────────────────────────────────────────────
for row_idx, (cat, code, desc, fee, notes) in enumerate(rows, start=2):
    fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
    values = [code, desc, fee if fee is not None else "—", cat, notes]
    aligns = [CENTER, LEFT, CENTER, LEFT, LEFT]
    for col_idx, (val, aln) in enumerate(zip(values, aligns), start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = CODE_FONT if col_idx == 1 else BODY_FONT
        cell.fill = fill
        cell.alignment = aln
        cell.border = BORDER
        if col_idx == 3 and isinstance(val, (int, float)):
            cell.number_format = '"$"#,##0'

ws.row_dimensions[1].height = 22

# ── add AutoFilter ────────────────────────────────────────────────────────────
ws.auto_filter.ref = f"A1:E{len(rows)+1}"

# ── second sheet: by category ────────────────────────────────────────────────
ws2 = wb.create_sheet("By Category")
ws2.freeze_panes = "A2"

for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = CENTER
    cell.border = hdr_border()
    ws2.column_dimensions[get_column_letter(col_idx)].width = w

ws2.row_dimensions[1].height = 22

from itertools import groupby
sorted_rows = sorted(rows, key=lambda r: r[0])
row_idx = 2
for cat_name, group in groupby(sorted_rows, key=lambda r: r[0]):
    group = list(group)
    # category header
    cat_cell = ws2.cell(row=row_idx, column=1, value=cat_name)
    cat_cell.font = CAT_FONT
    cat_cell.fill = CAT_FILL
    cat_cell.alignment = LEFT
    cat_cell.border = BORDER
    ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
    for c in range(2, 6):
        ws2.cell(row=row_idx, column=c).fill = CAT_FILL
        ws2.cell(row=row_idx, column=c).border = BORDER
    row_idx += 1
    for alt, (cat, code, desc, fee, notes) in enumerate(group):
        fill = ALT_FILL if alt % 2 == 0 else WHITE_FILL
        values = [code, desc, fee if fee is not None else "—", cat, notes]
        aligns = [CENTER, LEFT, CENTER, LEFT, LEFT]
        for col_idx, (val, aln) in enumerate(zip(values, aligns), start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = CODE_FONT if col_idx == 1 else BODY_FONT
            cell.fill = fill
            cell.alignment = aln
            cell.border = BORDER
            if col_idx == 3 and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0'
        row_idx += 1

ws2.auto_filter.ref = f"A1:E1"

out = r"C:\Users\chish\OneDrive\Desktop\Claude New Ohip\SMH_ED_OHIP_Billing_Apr2026.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total codes: {len(rows)}")
